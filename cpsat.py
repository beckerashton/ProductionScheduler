import collections
import json
import re
import stat
from asyncio import events
from datetime import date, timedelta
from math import trunc
from pathlib import Path
from typing import Dict, Optional, no_type_check

import matplotlib.pyplot as plt
import numpy as np
from matplotlib.pylab import f
from openpyxl import load_workbook
from ortools.sat.python import cp_model
from pandas import DataFrame
from pydantic import (BaseModel, Field, NonNegativeFloat, PositiveInt,
                      model_validator)
from pyexpat import model

from DbUtils import peek, refresh
from OtherUtils import tempShow
from Types import Event, EventGroup


# Util func to convert date to int for cpsat model relative to a fixed start date
# needed for compatibility with cp models constraints
def _date_to_int(this_date: date | str, start_date: date = date.today()) -> int:
    if isinstance(this_date, str):
        this_date = date.fromisoformat(this_date)
    return (this_date - start_date).days

# Container for the inputs to a scheduler instance
class SchedulerInstance(BaseModel):
    events: list[EventGroup] = Field(..., description="List of events to be scheduled.")
    machines: list[dict] = Field(..., description="List of machines available for scheduling.")

    @model_validator(mode="before")
    def validate_input(cls, data):
            # Basic validation to ensure required fields are present
        if "events" not in data or not isinstance(data["events"], list):
            raise ValueError("Input must contain a list of events.")
        if "machines" not in data or not isinstance(data["machines"], list):
            raise ValueError("Input must contain a list of machines.")
        return data

    @model_validator(mode="after")
    def reformat_events_dates(self):
        for event in self.events:
            if isinstance(event.requestedShipDate, str) or isinstance(event.requestedShipDate, date):
                event.requestedShipDate = _date_to_int(event.requestedShipDate) * 8 * 60
            if isinstance(event.estTime, float):
                event.estTime = int(event.estTime)
        return self

# Container for the configs of a scheduler instance
class SchedulerSolverConfig(BaseModel):
    time_limit_seconds: PositiveInt = Field(60, description="Time limit for the solver in seconds.")
    log_search_progress: bool = Field(False, description="Whether to log search progress during solving.")
    optimization_tolerance: NonNegativeFloat = Field(0.01, description="Tolerance for optimization.")
    num_search_workers: PositiveInt = Field(1, description="Number of parallel workers for the solver.")
    enumerate_all_solutions: bool = Field(False, description="Whether to enumerate all solutions (Must be equal to optimal).")


# Container for the outputs of a scheduler instance
class SchedulerSolution(BaseModel):
    schedule: list[dict] = Field(..., description="List of scheduled events with assigned machines and times.")
    objective_value: float = Field(..., description="Objective value of the solution.")
    status: str = Field(..., description="Status of the solver after attempting to solve the scheduling problem.")
    equally_optimal_schedules: list[list[dict]] = Field(
        default_factory=list,
        description="Additional schedules with the same optimal objective value.",
    )


def _event_duration(event: EventGroup) -> int:
    duration = getattr(event, "duration", None)
    if duration is None:
        duration = getattr(event, "estTime", None)
    if not isinstance(duration, int) or duration <= 0:
        print(event)
        raise ValueError(f"Event {event.groupId} must have a positive integer duration/estTime.")
    return duration


class _EventSchedulingVars:
    def __init__(self, instance: SchedulerInstance, model: cp_model.CpModel):
        self.event_to_machine = {}
        self.event_start = {}
        self.event_end = {}
        self.event_presence = {}
        self.machine_intervals = collections.defaultdict(list)

        self.horizon = sum(_event_duration(event) for event in instance.events) + max(event.requestedShipDate for event in instance.events)

        for event in instance.events:
            duration = _event_duration(event)
            requested_ship_date = event.requestedShipDate
            if not isinstance(requested_ship_date, int):
                raise ValueError(f"Event {event.groupId} requestedShipDate must be an integer for this prototype.")

            eligible_machine_ids = [m["id"] for m in instance.machines if m["colors"] >= event.colors and m["flashes"] >= event.flashes]
            # also manual add machine 4's secondary cap of 9/4 if 4 is not already in list
            if 4 not in eligible_machine_ids and event.colors <= 9 and event.flashes <= 4:
                eligible_machine_ids.append(4)

            if not eligible_machine_ids:
                raise ValueError(f"Event {event.groupId} has no eligible machines.")

            machine_var = model.new_int_var_from_domain(
                cp_model.Domain.FromValues(eligible_machine_ids),
                f"event_{event.groupId}_machine"
            )
            start_var = model.new_int_var(0, self.horizon - duration, f"event_{event.groupId}_start")
            end_var = model.new_int_var(duration, self.horizon, f"event_{event.groupId}_end")

            presence_vars = []
            for machine_id in eligible_machine_ids:
                presence = model.new_bool_var(f"event_{event.groupId}_on_machine_{machine_id}")
                interval_var = model.new_optional_interval_var(
                    start_var,
                    duration,
                    end_var,
                    presence,
                    f"event_{event.groupId}_interval_machine_{machine_id}",
                )
                self.event_presence[event.groupId, machine_id] = presence
                self.machine_intervals[machine_id].append(interval_var)
                model.add(machine_var == machine_id).only_enforce_if(presence)
                presence_vars.append(presence)

            model.add_exactly_one(presence_vars)

            self.event_to_machine[event.groupId] = machine_var
            self.event_start[event.groupId] = start_var
            self.event_end[event.groupId] = end_var

@no_type_check # this func is goofy 
def _read_solution_value(value_source: object, expression) -> int:
    value_fn = getattr(value_source, "value", None)
    if callable(value_fn):
        return int(value_fn(expression))
    return int(value_source.Value(expression))


def _build_schedule_snapshot(
    instance: SchedulerInstance,
    event_to_machine: Dict[int, cp_model.IntVar],
    event_start: Dict[int, cp_model.IntVar],
    event_end: Dict[int, cp_model.IntVar],
    value_source: object,
) -> list[dict]:
    return [
        {
            "groupId": event.groupId,
            "designId": event.designId,
            "assignedMachineId": _read_solution_value(value_source, event_to_machine[event.groupId]),
            "scheduledStartDate": _read_solution_value(value_source, event_start[event.groupId]),
            "scheduledEndDate": _read_solution_value(value_source, event_end[event.groupId]),
            "requestedShipDate": event.requestedShipDate,
            "colors": event.colors,
            "flashes": event.flashes,
        }
        for event in instance.events
    ]


def _schedule_signature(schedule: list[dict]) -> tuple[tuple[int, int, int, int], ...]:
    return tuple(
        (job["groupId"], job["assignedMachineId"], job["scheduledStartDate"], job["scheduledEndDate"])
        for job in sorted(schedule, key=lambda row: row["groupId"])
    )


class _ScheduleCollector(cp_model.CpSolverSolutionCallback):
    def __init__(
        self,
        instance: SchedulerInstance,
        event_to_machine: Dict[int, cp_model.IntVar],
        event_start: Dict[int, cp_model.IntVar],
        event_end: Dict[int, cp_model.IntVar],
    ):
        super().__init__()
        self._instance = instance
        self._event_to_machine = event_to_machine
        self._event_start = event_start
        self._event_end = event_end
        self._seen_signatures: set[tuple[tuple[int, int, int, int], ...]] = set()
        self.schedules: list[list[dict]] = []

    def on_solution_callback(self) -> None:
        schedule = _build_schedule_snapshot(
            self._instance,
            self._event_to_machine,
            self._event_start,
            self._event_end,
            self,
        )
        schedule_signature = _schedule_signature(schedule)
        if schedule_signature in self._seen_signatures:
            return
        self._seen_signatures.add(schedule_signature)
        self.schedules.append(schedule)


class SchedulerSolver:
    def __init__(self, instance: SchedulerInstance, config: SchedulerSolverConfig):
        self.instance = instance
        self.config = config
        self.model = cp_model.CpModel()
        self._event_vars = _EventSchedulingVars(instance, self.model)
        self._objective_var: cp_model.IntVar | None = None
        self._build_model()
        self.solver = cp_model.CpSolver()

    def _add_default_constraints(self):
        # self._add_constraint_force_before_ship_date()
        self._add_constraint_machine_no_overlap()
        self._add_constraint_machine_contiguous_block()
        # self._add_constraint_pad_between_events()

    def _add_constraint_force_before_ship_date(self):
        for event in self.instance.events:
            self.model.add(self._event_vars.event_end[event.groupId] <= event.requestedShipDate)

    def _add_constraint_force_before_ship_date_ignore_lates(self):
        for event in self.instance.events:
            if event.requestedShipDate > 0:  # Only enforce for events that are not already late
                self.model.add(self._event_vars.event_end[event.groupId] <= event.requestedShipDate)

    # [Not updated since estTime split, complexity split, and requestedShipDate int conversion, may need adjustments to work with new event structure]
    # def _add_constraint_force_before_ship_date_ignore_hinted(self, pre_solution: Optional[SchedulerSolution] = None):
    #     hinted_event_ids = set()
    #     if pre_solution:
    #         hinted_event_ids = {e["groupId"] for e in pre_solution.schedule}
        
    #     for event in self.instance.events:
    #         if event.requestedShipDate > 0 and event.groupId not in hinted_event_ids:
    #             self.model.add(self._event_vars.event_end[event.groupId] <= event.requestedShipDate)

    def _add_constraint_machine_no_overlap(self):
        for machine in self.instance.machines:
            machine_id = machine["id"]
            intervals_on_machine = self._event_vars.machine_intervals[machine_id]
            if intervals_on_machine:
                self.model.add_no_overlap(intervals_on_machine)

    def _add_constraint_machine_contiguous_block(self):
        horizon = self._event_vars.horizon

        for machine in self.instance.machines:
            machine_id = machine["id"]
            machine_events = [
                event
                for event in self.instance.events
                if (event.groupId, machine_id) in self._event_vars.event_presence
            ]
            if not machine_events:
                continue

            presences = [
                self._event_vars.event_presence[event.groupId, machine_id]
                for event in machine_events
            ]

            machine_used = self.model.new_bool_var(f"machine_{machine_id}_used")
            self.model.add(sum(presences) >= 1).only_enforce_if(machine_used)
            self.model.add(sum(presences) == 0).only_enforce_if(machine_used.Not())

            adjusted_starts = []
            adjusted_ends = []

            for event in machine_events:
                presence = self._event_vars.event_presence[event.groupId, machine_id]
                start = self._event_vars.event_start[event.groupId]
                end = self._event_vars.event_end[event.groupId]

                adjusted_start = self.model.new_int_var(
                    0,
                    horizon,
                    f"event_{event.groupId}_adjusted_start_machine_{machine_id}",
                )
                adjusted_end = self.model.new_int_var(
                    0,
                    horizon,
                    f"event_{event.groupId}_adjusted_end_machine_{machine_id}",
                )

                self.model.add(adjusted_start == start).only_enforce_if(presence)
                self.model.add(adjusted_start == horizon).only_enforce_if(presence.Not())

                self.model.add(adjusted_end == end).only_enforce_if(presence)
                self.model.add(adjusted_end == 0).only_enforce_if(presence.Not())

                adjusted_starts.append(adjusted_start)
                adjusted_ends.append(adjusted_end)

            first_start = self.model.new_int_var(0, horizon, f"machine_{machine_id}_first_start")
            last_end = self.model.new_int_var(0, horizon, f"machine_{machine_id}_last_end")
            self.model.add_min_equality(first_start, adjusted_starts)
            self.model.add_max_equality(last_end, adjusted_ends)
            self.model.add(first_start == 0).only_enforce_if(machine_used)

            busy_span = self.model.new_int_var(0, horizon, f"machine_{machine_id}_busy_span")
            self.model.add(busy_span == last_end - first_start).only_enforce_if(machine_used)
            self.model.add(busy_span == 0).only_enforce_if(machine_used.Not())

            total_processing_time = sum(
                _event_duration(event) * self._event_vars.event_presence[event.groupId, machine_id]
                for event in machine_events
            )
            self.model.add(total_processing_time == busy_span)

    def _add_constraint_sequence_subevents(self):
        # force events with the same designId root to have its subevents scheduled in misc>Front Left Chest>Sleeve>Full Front>Full Back order
        root_to_subevents = collections.defaultdict(list)
        for event in self.instance.events:
            root_id = event.designId.split("_")[0]
            root_to_subevents[root_id].append(event)
        for root_id, subevents in root_to_subevents.items():
            if len(subevents) <= 1:
                continue
            # sort order Front Left Chest>Sleeve>Full Front>Full Back with any unspecified going first
            subevents.sort(key=lambda e: ["Front Left Chest", "Sleeve", "Full Front", "Full Back"].index(e.designId.split("_")[1]) if e.designId.split("_")[1] in ["Front Left Chest", "Sleeve", "Full Front", "Full Back"] else -1)
            for i in range(len(subevents) - 1):
                event_i = subevents[i]
                event_j = subevents[i + 1]
                self.model.add(self._event_vars.event_end[event_i.groupId] <= self._event_vars.event_start[event_j.groupId])

    # constraint that on the same machine there is a gap of 1 time unit between events
    def _add_constraint_pad_between_events(self):
        for machine in self.instance.machines:
            machine_id = machine["id"]
            intervals_on_machine = self._event_vars.machine_intervals[machine_id]
            if intervals_on_machine:
                for i in range(len(intervals_on_machine)):
                    for j in range(i + 1, len(intervals_on_machine)):
                        interval_i = intervals_on_machine[i]
                        interval_j = intervals_on_machine[j]
                        # Extract groupId from interval name: "event_{groupId}_interval_machine_{machineId}"
                        event_id_i = int(interval_i.Name().split("_")[1])
                        event_id_j = int(interval_j.Name().split("_")[1])
                        presence_i = self._event_vars.event_presence[event_id_i, machine_id]
                        presence_j = self._event_vars.event_presence[event_id_j, machine_id]
                        
                        # Add padding constraint only when both events are on this machine
                        self.model.add(interval_i.EndExpr() + 1 <= interval_j.StartExpr()).only_enforce_if(
                            [presence_i, presence_j]
                        )
                        self.model.add(interval_j.EndExpr() + 1 <= interval_i.StartExpr()).only_enforce_if(
                            [presence_i, presence_j]
                        )
                        
    def _add_soft_deadline_penalty(self):
        penalties = []
        for event in self.instance.events:
            tardiness = self.model.new_int_var(0, 10000, f"tardiness_{event.groupId}")
            self.model.add_max_equality(tardiness, [
                self._event_vars.event_end[event.groupId] - event.requestedShipDate,
                0
            ])
            penalties.append(tardiness)
        return penalties
    
    def _add_presolve_hint(self, pre_solution: SchedulerSolution):
        for scheduled_event in pre_solution.schedule:
            order_id = scheduled_event["groupId"]
            assigned_machine_id = scheduled_event["assignedMachineId"]
            scheduled_start = scheduled_event["scheduledStartDate"]
            self.model.add(self._event_vars.event_to_machine[order_id] == assigned_machine_id)
            self.model.add(self._event_vars.event_start[order_id] == scheduled_start)

    def _set_makespan_objective(self):
        makespan = self.model.new_int_var(0, self._event_vars.horizon, "makespan")
        self.model.add_max_equality(makespan, [self._event_vars.event_end[event.groupId] for event in self.instance.events])
        self._objective_var = makespan
        self.model.minimize(makespan)

    # Multi-layered makespan objective that minimizes the makespan on each subset of machines
    # where each iteration of ignores the previous max makespan machines to create a secondary objective of minimizing the makespan of the remaining machines, and so on for a specified number of iterations (makespan_checks)
    def _set_multi_makespan_objective(self, makespan_checks: int):
        if makespan_checks <= 0:
            raise ValueError("makespan_checks must be >= 1")

        remaining_machine_ids = [machine["id"] for machine in self.instance.machines]
        objective_terms = []

        for check in range(makespan_checks):
            makespan = self.model.new_int_var(0, self._event_vars.horizon, f"makespan_check_{check}")
            adjusted_event_ends = []

            for event in self.instance.events:
                presences_on_remaining = [
                    self._event_vars.event_presence[event.groupId, machine_id]
                    for machine_id in remaining_machine_ids
                    if (event.groupId, machine_id) in self._event_vars.event_presence
                ]
                if not presences_on_remaining:
                    continue

                if len(presences_on_remaining) == 1:
                    event_on_remaining = presences_on_remaining[0]
                else:
                    event_on_remaining = self.model.new_bool_var(
                        f"event_{event.groupId}_on_remaining_check_{check}"
                    )
                    # Exactly-one assignment makes this sum either 0 or 1.
                    self.model.add(event_on_remaining == sum(presences_on_remaining))

                adjusted_end = self.model.new_int_var(
                    0,
                    self._event_vars.horizon,
                    f"event_{event.groupId}_adjusted_end_check_{check}",
                )
                self.model.add(adjusted_end == self._event_vars.event_end[event.groupId]).only_enforce_if(
                    event_on_remaining
                )
                self.model.add(adjusted_end == 0).only_enforce_if(event_on_remaining.Not())
                adjusted_event_ends.append(adjusted_end)

            if adjusted_event_ends:
                self.model.add_max_equality(makespan, adjusted_event_ends)
            else:
                self.model.add(makespan == 0)

            objective_terms.append(makespan * (makespan_checks - check))

            if check < makespan_checks - 1 and remaining_machine_ids:
                # Decision variables cannot be sorted in Python, so peel off a deterministic
                # prefix of machines for each subsequent tier.
                remove_count = max(1, len(remaining_machine_ids) // makespan_checks)
                remaining_machine_ids = remaining_machine_ids[remove_count:]

        weighted_objective_upper_bound = self._event_vars.horizon * sum(
            makespan_checks - check for check in range(len(objective_terms))
        )
        weighted_objective = self.model.new_int_var(
            0,
            weighted_objective_upper_bound,
            "multi_makespan_objective",
        )
        self.model.add(weighted_objective == sum(objective_terms))
        self._objective_var = weighted_objective
        self.model.minimize(weighted_objective)

    def _set_makespan_with_tardiness_penalty_objective(self):
        makespan = self.model.new_int_var(0, self._event_vars.horizon, "makespan")
        self.model.add_max_equality(makespan, [self._event_vars.event_end[event.groupId] for event in self.instance.events])
        penalties = self._add_soft_deadline_penalty()
        max_total_penalty = 10000 * len(self.instance.events)
        weighted_objective = self.model.new_int_var(
            0,
            self._event_vars.horizon + 1000 * max_total_penalty,
            "makespan_with_tardiness_penalty",
        )
        self.model.add(weighted_objective == makespan + 1000 * sum(penalties))
        self._objective_var = weighted_objective
        self.model.minimize(weighted_objective)

    def _set_balanced_objective(self):
        """
        [WARNING] IGNORES STRYKER
        """
        # Minimize makespan AND variance in machine load
        makespan = self.model.new_int_var(0, self._event_vars.horizon, "makespan")
        self.model.add_max_equality(makespan, [self._event_vars.event_end[e.groupId] for e in self.instance.events])
        
        machine_loads = []
        for machine in self.instance.machines:
            if machine["id"] == 6:  # ignores stryker when balancing loads
                continue
            load = sum(
                self._event_vars.event_presence[event.groupId, machine["id"]] * event.estTime
                for event in self.instance.events
                # if machine["colors"] >= event.colors and machine["flashes"] >= event.flashes # i think this is redundant
            )
            machine_loads.append(load)
        
        # Penalize maximum load difference
        max_load = self.model.new_int_var(0, self._event_vars.horizon, "max_load")
        min_load = self.model.new_int_var(0, self._event_vars.horizon, "min_load")
        self.model.add_max_equality(max_load, machine_loads)
        self.model.add_min_equality(min_load, machine_loads)

        balanced_objective = self.model.new_int_var(0, self._event_vars.horizon * 11, "balanced_objective")
        self.model.add(balanced_objective == makespan * 10 + (max_load - min_load))
        self._objective_var = balanced_objective
        self.model.minimize(balanced_objective)  # Weighted multi-objective

    def _build_model(self):
        self._add_default_constraints()

    def _configure_solver(
        self,
        solver: cp_model.CpSolver,
        time_limit: float | None,
        *,
        enumerate_all_solutions: bool = False,
        force_exact_optimal: bool = False,
        num_search_workers: int | None = None,
    ) -> None:
        solver.parameters.max_time_in_seconds = time_limit if time_limit is not None else self.config.time_limit_seconds
        solver.parameters.log_search_progress = self.config.log_search_progress
        solver.parameters.relative_gap_limit = 0.0 if force_exact_optimal else self.config.optimization_tolerance
        solver.parameters.num_search_workers = num_search_workers if num_search_workers is not None else self.config.num_search_workers
        solver.parameters.enumerate_all_solutions = enumerate_all_solutions

    def _enumerate_equally_optimal_schedules(
        self,
        primary_schedule: list[dict],
        optimal_objective_value: float,
        time_limit: float | None = None,
    ) -> list[list[dict]]:
        if self._objective_var is None:
            print("Skipped equally optimal enumeration because no objective has been set on the model.")
            return []

        optimal_objective = int(round(optimal_objective_value))
        enumeration_model = self.model.clone()
        enumeration_objective_var = enumeration_model.get_int_var_from_proto_index(self._objective_var.Index())
        enumeration_model.clear_objective()
        enumeration_model.add(enumeration_objective_var == optimal_objective)

        cloned_event_to_machine: Dict[int, cp_model.IntVar] = {
            group_id: enumeration_model.get_int_var_from_proto_index(var.Index())
            for group_id, var in self._event_vars.event_to_machine.items()
        }
        cloned_event_start: Dict[int, cp_model.IntVar] = {
            group_id: enumeration_model.get_int_var_from_proto_index(var.Index())
            for group_id, var in self._event_vars.event_start.items()
        }
        cloned_event_end: Dict[int, cp_model.IntVar] = {
            group_id: enumeration_model.get_int_var_from_proto_index(var.Index())
            for group_id, var in self._event_vars.event_end.items()
        }

        collector = _ScheduleCollector(
            self.instance,
            cloned_event_to_machine,
            cloned_event_start,
            cloned_event_end,
        )
        enumeration_solver = cp_model.CpSolver()
        self._configure_solver(
            enumeration_solver,
            time_limit,
            enumerate_all_solutions=True,
            force_exact_optimal=True,
            num_search_workers=1,
        )
        enumeration_status = enumeration_solver.SearchForAllSolutions(enumeration_model, collector)

        primary_signature = _schedule_signature(primary_schedule)
        alternate_schedules = [
            schedule
            for schedule in collector.schedules
            if _schedule_signature(schedule) != primary_signature
        ]

        if enumeration_status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
            print(
                "Equal-optimum enumeration failed "
                f"(status={enumeration_solver.StatusName(enumeration_status)}); "
                "returning the primary optimal solution only."
            )
            return []

        if enumeration_status == cp_model.FEASIBLE:
            print(
                "Equal-optimum enumeration hit the solver limit before exhausting all schedules; "
                f"returning {len(alternate_schedules)} alternate schedule(s) found so far."
            )

        return alternate_schedules

    def solve(self, time_limit: float | None = None) -> SchedulerSolution:
        self._configure_solver(
            self.solver,
            time_limit,
            enumerate_all_solutions=False,
            force_exact_optimal=self.config.enumerate_all_solutions,
        )
        status = self.solver.Solve(self.model)

        if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
            solution = SchedulerSolution(
                schedule=_build_schedule_snapshot(
                    self.instance,
                    self._event_vars.event_to_machine,
                    self._event_vars.event_start,
                    self._event_vars.event_end,
                    self.solver,
                ),
                objective_value=self.solver.ObjectiveValue(),
                status=self.solver.StatusName(status),
            )

            if self.config.enumerate_all_solutions:
                if status != cp_model.OPTIMAL:
                    print("Skipped equally optimal enumeration because the solver did not prove optimality.")
                else:
                    solution.equally_optimal_schedules = self._enumerate_equally_optimal_schedules(
                        solution.schedule,
                        solution.objective_value,
                        time_limit,
                    )

            return solution

        return SchedulerSolution(
            schedule=[],
            objective_value=0.0,
            status=self.solver.StatusName(status)
        )


def _write_excel_cells_with_app(filename: str, sheet_name: str, cell_updates: list[tuple[int, int, int]]) -> str:
    try:
        from win32com.client import DispatchEx
    except ImportError as exc:
        raise RuntimeError(
            "Writing Excel files while preserving external links requires pywin32."
        ) from exc

    source_path = Path(filename).resolve()
    output_path = source_path.with_name(f"{source_path.stem}_with_schedule({date.today().day}){source_path.suffix}")

    excel = DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False

    workbook = None
    worksheet = None
    try:
        workbook = excel.Workbooks.Open(str(source_path), UpdateLinks=0, ReadOnly=False)
        worksheet = workbook.Worksheets(sheet_name)

        for row_idx, column_idx, value in cell_updates:
            worksheet.Cells(row_idx, column_idx).Value = value

        workbook.SaveCopyAs(str(output_path))
        return str(output_path)
    finally:
        if workbook is not None:
            workbook.Close(SaveChanges=False)
        worksheet = None
        excel.Quit()


def _save_interactive_schedule_graph(graph_rows: list[dict], workday_minutes: int, output_path: str | Path) -> str | None:
    if not graph_rows:
        return None

    try:
        import plotly.graph_objects as go
    except ImportError as exc:
        raise RuntimeError(
            "Saving an interactive schedule graph requires plotly."
        ) from exc

    resolved_output_path = Path(output_path).resolve()
    resolved_output_path.parent.mkdir(parents=True, exist_ok=True)

    location_labels = sorted({row["location"] for row in graph_rows})
    palette = [
        "#1b9e77",
        "#d95f02",
        "#7570b3",
        "#e7298a",
        "#66a61e",
        "#e6ab02",
        "#a6761d",
        "#666666",
        "#0b84a5",
        "#f6c85f",
        "#ca472f",
        "#8dddd0",
        "#b30000",
    ]
    location_to_color = {
        location: palette[index % len(palette)]
        for index, location in enumerate(location_labels)
    }

    fig = go.Figure()
    legend_seen = set()

    for row in graph_rows:
        location = row["location"]
        show_legend = location not in legend_seen
        legend_seen.add(location)

        fig.add_trace(
            go.Bar(
                x=[row["duration"]],
                y=[row["machineLabel"]],
                base=[row["start"]],
                orientation="h",
                marker={
                    "color": location_to_color[location],
                    "line": {"color": "black", "width": 0.5},
                },
                text=[str(row["groupId"])],
                textposition="inside",
                name=location,
                legendgroup=f"location_{location}",
                showlegend=show_legend,
                customdata=[[
                    row["groupId"],
                    row["designId"],
                    row["machineId"],
                    row["startText"],
                    row["endText"],
                    row["duration"],
                    row["estTime"],
                    row["requestedShipText"],
                    location,
                ]],
                hovertemplate=(
                    "Order ID: %{customdata[0]}<br>"
                    "Design ID: %{customdata[1]}<br>"
                    "Location: %{customdata[8]}<br>"
                    "Machine ID: %{customdata[2]}<br>"
                    "Start: %{customdata[3]}<br>"
                    "End: %{customdata[4]}<br>"
                    "Duration: %{customdata[5]}<br>"
                    "Est Time: %{customdata[6]}<br>"
                    "Requested Ship: %{customdata[7]}"
                    "<extra></extra>"
                ),
            )
        )

    max_end = max(row["end"] for row in graph_rows)
    tick_limit = ((max_end + workday_minutes - 1) // workday_minutes + 1) * workday_minutes
    tick_values = list(range(0, tick_limit + 1, workday_minutes))
    tick_labels = [str(value // workday_minutes) for value in tick_values]
    machine_labels = sorted({row["machineLabel"] for row in graph_rows}, key=lambda label: int(label.split()[-1]))

    fig.update_layout(
        title="Schedule by Machine (interactive)",
        template="plotly_white",
        hovermode="closest",
        barmode="overlay",
        xaxis={
            "title": "Day",
            "tickmode": "array",
            "tickvals": tick_values,
            "ticktext": tick_labels,
            "showgrid": True,
            "gridcolor": "rgba(0, 0, 0, 0.15)",
        },
        yaxis={
            "title": "Machine",
            "categoryorder": "array",
            "categoryarray": machine_labels,
        },
        legend={"title": {"text": "Location"}},
    )

    fig.write_html(
        str(resolved_output_path),
        include_plotlyjs=True,
        full_html=True,
        config={"displaylogo": False},
    )
    return str(resolved_output_path)

    
def write_event_sequence_into_excel(solution: SchedulerSolution, filename: str) -> None:
    
    # Read values with openpyxl, but write/save through Excel so workbook links survive.
    wb = load_workbook(filename, data_only=True, keep_links=True)
    ws = wb["Sheet3"]
    
    # Create a mapping of (Design No root, Location) -> solution data
    solution_map = {}
    for scheduled_event in solution.schedule:
        design_id: str = scheduled_event["designId"]
        design_root, location = design_id.split("_", 1)
        solution_map[(design_root, location)] = {
            "assignedMachine": scheduled_event["assignedMachineId"],
            "scheduledStart": scheduled_event["scheduledStartDate"],
            "setupCompleted": False
        }
    
    # remap scheduledStart to instead be the order per machine (1, 2, 3...)
    machine_to_events = collections.defaultdict(list)
    for scheduled_event in solution.schedule:
        machine_to_events[scheduled_event["assignedMachineId"]].append(scheduled_event)
    for machine_id, events in machine_to_events.items():
        events.sort(key=lambda e: e["scheduledStartDate"])
        for sequence, event in enumerate(events, start=1):
            design_id = event["designId"]
            design_root, location = design_id.split("_", 1)
            solution_map[(design_root, location)]["sequence"] = sequence
    
    cell_updates = []

    # Iterate through rows starting from header row + 1
    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        # Assuming Design No is in column C (index 2) and Location is in column E (index 4)
        design_no_cell = ws.cell(row=row_idx, column=3)  # Column C
        location_cell = ws.cell(row=row_idx, column=5)   # Column E

        design_no = design_no_cell.value
        location = location_cell.value
        
        if design_no is None or location is None:
            continue
        
        design_root = str(design_no).split(".", 1)[0].strip()
        location_str = str(location).strip()
        
        # Look up in solution map
        if (design_root, location_str) in solution_map:
            scheduled_data = solution_map[(design_root, location_str)]

            # Write to Press column (assuming column R, index 18) and Sequence column (assuming column S, index 19)
            cell_updates.append((row_idx, 18, scheduled_data["assignedMachine"]))
            cell_updates.append((row_idx, 19, scheduled_data["sequence"]))
            if scheduled_data["setupCompleted"] == True:
                cell_updates.append((row_idx, 20, "1"))
            else:
                solution_map[(design_root, location_str)]["setupCompleted"] = True

    wb.close()
    _write_excel_cells_with_app(filename, "Sheet3", cell_updates)


def _model_minutes_to_datetime_text(model_minutes: float, workday_minutes: int) -> str:
    clamped_minutes = max(0, int(round(model_minutes)))
    day_offset, minute_of_day = divmod(clamped_minutes, workday_minutes)
    actual_date = date.today() + timedelta(days=day_offset)
    hour = 8 + (minute_of_day // 60)
    minute = minute_of_day % 60
    return f"{actual_date.isoformat()} {hour:02d}:{minute:02d}"


def _design_location_from_design_id(design_id: str) -> str:
    if "_" not in design_id:
        return "Unknown"
    _design_root, location = design_id.split("_", 1)
    normalized_location = location.strip()
    return normalized_location if normalized_location else "Unknown"


def _plot_schedule_graph(
    schedule: list[dict],
    instance: SchedulerInstance,
    workday_minutes: int,
    title: str,
    interactive_output_path: str | Path | None = None,
) -> str | None:
    from matplotlib.patches import Rectangle
    from matplotlib.ticker import FuncFormatter, MultipleLocator

    fig, ax = plt.subplots(figsize=(14, 8))
    event_by_order_id = {event.groupId: event for event in instance.events}

    location_labels = sorted({
        _design_location_from_design_id(event_by_order_id[job["groupId"]].designId)
        for job in schedule
    })
    colors = plt.cm.get_cmap("tab20")(np.linspace(0, 1, max(1, len(location_labels))))
    location_to_color = {
        location: colors[index]
        for index, location in enumerate(location_labels)
    }

    bars_with_details = []
    interactive_schedule_rows = []

    machine_schedules = collections.defaultdict(list)
    for job in schedule:
        machine_schedules[job["assignedMachineId"]].append(job)

    yticks = []
    ylabels = []
    y_pos = 0

    for machine_id in sorted(machine_schedules.keys()):
        jobs = sorted(machine_schedules[machine_id], key=lambda job: job["scheduledStartDate"])
        yticks.append(y_pos)
        ylabels.append(f"Machine {machine_id}")

        for job in jobs:
            event = event_by_order_id[job["groupId"]]
            location = _design_location_from_design_id(event.designId)
            color = location_to_color[location]

            bar_container = ax.barh(
                y_pos,
                job["scheduledEndDate"] - job["scheduledStartDate"],
                left=job["scheduledStartDate"],
                height=0.6,
                color=color,
                edgecolor="black",
                linewidth=0.5,
            )
            bar = bar_container.patches[0]
            details_text = (
                f"Order ID: {job['groupId']}\n"
                f"Design ID: {event.designId}\n"
                f"Location: {location}\n"
                f"Machine ID: {job['assignedMachineId']}\n"
                f"Start: {_model_minutes_to_datetime_text(job['scheduledStartDate'], workday_minutes)}\n"
                f"End: {_model_minutes_to_datetime_text(job['scheduledEndDate'], workday_minutes)}\n"
                f"Duration: {job['scheduledEndDate'] - job['scheduledStartDate']}\n"
                f"Est Time: {event.estTime}\n"
                f"Requested Ship: {_model_minutes_to_datetime_text(event.requestedShipDate, workday_minutes)}\n"
                f"Colors: {event.colors}, Flashes: {event.flashes}"
            )
            bars_with_details.append((bar, details_text))
            interactive_schedule_rows.append({
                "groupId": job["groupId"],
                "designId": event.designId,
                "location": location,
                "machineId": job["assignedMachineId"],
                "machineLabel": f"Machine {job['assignedMachineId']}",
                "start": job["scheduledStartDate"],
                "end": job["scheduledEndDate"],
                "duration": job["scheduledEndDate"] - job["scheduledStartDate"],
                "startText": _model_minutes_to_datetime_text(job["scheduledStartDate"], workday_minutes),
                "endText": _model_minutes_to_datetime_text(job["scheduledEndDate"], workday_minutes),
                "estTime": event.estTime,
                "requestedShipText": _model_minutes_to_datetime_text(event.requestedShipDate, workday_minutes),
                "colors": event.colors,
                "flashes": event.flashes,
            })
            ax.text(
                job["scheduledStartDate"] + (job["scheduledEndDate"] - job["scheduledStartDate"]) / 2,
                y_pos,
                f"{job['groupId']}",
                ha="center",
                va="center",
                fontsize=8,
                color="white",
                weight="bold",
            )

        y_pos += 1

    ax.set_yticks(yticks)
    ax.set_yticklabels(ylabels)
    ax.xaxis.set_major_locator(MultipleLocator(workday_minutes))
    ax.xaxis.set_major_formatter(FuncFormatter(lambda value, pos: value // workday_minutes))
    ax.set_ylabel("Machine")
    ax.set_title(title)
    ax.grid(axis="x", alpha=0.3)

    legend_elements = [
        Rectangle((0, 0), 1, 1, facecolor=location_to_color[location], edgecolor="black", label=location)
        for location in location_labels
    ]
    if legend_elements:
        ax.legend(handles=legend_elements, loc="upper right", title="Location")

    tooltip = ax.annotate(
        "",
        xy=(0, 0),
        xytext=(10, 10),
        textcoords="offset points",
        bbox={"boxstyle": "round,pad=0.4", "fc": "white", "ec": "black", "alpha": 0.9},
        arrowprops={"arrowstyle": "->", "color": "black"},
    )
    tooltip.set_visible(False)

    def _update_tooltip(selected_bar, details_text: str):
        tooltip.xy = (
            selected_bar.get_x() + selected_bar.get_width() / 2,
            selected_bar.get_y() + selected_bar.get_height() / 2,
        )
        tooltip.set_text(details_text)

    def _on_hover(mouse_event):
        if mouse_event.inaxes != ax:
            if tooltip.get_visible():
                tooltip.set_visible(False)
                fig.canvas.draw_idle()
            return

        for bar, details_text in bars_with_details:
            contains, _ = bar.contains(mouse_event)
            if contains:
                _update_tooltip(bar, details_text)
                if not tooltip.get_visible():
                    tooltip.set_visible(True)
                fig.canvas.draw_idle()
                return

        if tooltip.get_visible():
            tooltip.set_visible(False)
            fig.canvas.draw_idle()

    fig.canvas.mpl_connect("motion_notify_event", _on_hover)
    fig.tight_layout()

    if interactive_output_path is None:
        return None

    interactive_graph_path = _save_interactive_schedule_graph(
        interactive_schedule_rows,
        workday_minutes,
        interactive_output_path,
    )
    if interactive_graph_path:
        print(f"Saved interactive schedule graph to {interactive_graph_path}")
    return interactive_graph_path


def main(show_graph: bool = False, save_graph: bool = False, write_solution_to_excel: bool = False):
    import pandas as pd
    excel_path: str = "C:/Users/Aston/Documents/Production Scheduler Demo.xlsx"
    
    input_df = pd.read_excel(excel_path, sheet_name="Sheet3", header=1, usecols="B,C,E,M,P,U,V,W")
    input_df = input_df.dropna(subset=["Order No", "Design No", "Location", "DueDate", "Imp", "No_Colors", "No_Flashes"])
    
    df = input_df.rename(columns={"Order No": "id_Order", "Design No": "id_Design", "Location": "Location", "DueDate": "date_OrderRequestedToShip", "Imp": "cn_QtyToProduce", "No_Colors": "ColorsTotal", "No_Flashes": "flashes"})
    df = df.dropna(subset=["id_Order", "id_Design", "Location", "date_OrderRequestedToShip", "cn_QtyToProduce", "ColorsTotal", "flashes"])

    df["runTime"] = df.apply(lambda row: row["cn_QtyToProduce"] / 250 * 60, axis=1)
    df["setupTime"] = df.apply(lambda row: row["ColorsTotal"] * 10, axis=1)
    df["colors"] = df["ColorsTotal"]
    df["flashes"] = df["flashes"]
    df["designId"] = df.apply(lambda row: f"{int(row['id_Design'])}_{row['Location']}", axis=1)
    df["date_OrderRequestedToShip"] = df["date_OrderRequestedToShip"].apply(lambda x: x.date() if isinstance(x, pd.Timestamp) else date.fromisoformat(x) if isinstance(x, str) else x)

    events = [
        Event(
            row["id_Order"], row["designId"], 
            row["runTime"], row["setupTime"], 
            row["date_OrderRequestedToShip"], 
            row["colors"], row["flashes"]
        ) for _, row in df.iterrows()
    ]

    events_grouped = collections.defaultdict(lambda: {"estTime": 0, "requestedShipDate": 0, "colors": 0, "flashes": 0})
    for event in events:
        if event.designId not in events_grouped:
            events_grouped[event.designId] = {"estTime": event.setupTime + event.runTime, "requestedShipDate": event.requestedShipDate, "colors": event.colors, "flashes": event.flashes} # pyright: ignore[reportArgumentType]
        else:
            events_grouped[event.designId]["estTime"] += event.runTime
            events_grouped[event.designId]["requestedShipDate"] = min(events_grouped[event.designId]["requestedShipDate"], event.requestedShipDate) # pyright: ignore[reportArgumentType]
            events_grouped[event.designId]["colors"] = max(events_grouped[event.designId]["colors"], event.colors)
            events_grouped[event.designId]["flashes"] = max(events_grouped[event.designId]["flashes"], event.flashes)
    
    max_presolve_window_days = 7
    instance: Optional[SchedulerInstance] = None
    solution: Optional[SchedulerSolution] = None
    for presolve_days in range(0, max_presolve_window_days + 1, 1):
        # late_events_grouped = {k: v for k, v in events_grouped.items() if date.fromisoformat(v["requestedShipDate"]) <= date.today() + pd.Timedelta(days=presolve_days)}

        machines = [
            {"id": 1, "colors": 12, "flashes": 3},
            {"id": 2, "colors": 8, "flashes": 3},
            {"id": 3, "colors": 12, "flashes": 3},
            {"id": 4, "colors": 12, "flashes": 3},  # override for 9/4 designs
            {"id": 5, "colors": 6, "flashes": 2},
            {"id": 6, "colors": 50, "flashes": 10},
            {"id": 7, "colors": 6, "flashes": 3},
        ]

        # pre solve with tighter constraints on late events to use for a min solve hint for a full solve
        # pre_machines = [v for v in machines if v["capability"] <= max(event["complexity"] for event in late_events_grouped.values())]

        # pre_instance = SchedulerInstance(events=[DummyEvent(i, k, v["estTime"], v["requestedShipDate"], v["complexity"]) for i, (k, v) in enumerate(late_events_grouped.items())],
        #     machines=pre_machines
        # )

        config = SchedulerSolverConfig(
            time_limit_seconds=30, 
            log_search_progress=False, 
            optimization_tolerance=0.01, 
            num_search_workers=16,
            enumerate_all_solutions=False
        )
        # pre_solver = SchedulerSolver(pre_instance, config)
        # pre_solver._set_balanced_objective()
        # pre_solver._add_constraint_sequence_subevents()
        # pre_solution = pre_solver.solve(time_limit=30)
        # print(f"Pre-solve on late events - Solution status: {pre_solution.status}, Objective value: {pre_solution.objective_value}")

        # save solution for debugging
        # pre_solution_df = pd.DataFrame(pre_solution.schedule)
        # pre_solution_df.to_csv("Inputs/pre_solution.csv", index=False)

        instance = SchedulerInstance(events=[EventGroup(i, k, v["estTime"], v["colors"], v["flashes"], v["requestedShipDate"]) for i, (k, v) in enumerate(events_grouped.items())],
            machines=machines
        )
        
        solver = SchedulerSolver(instance, config)
        # solver._add_presolve_hint(pre_solution)
        # solver._set_makespan_objective()
        solver._set_multi_makespan_objective(makespan_checks=3)
        # solver._add_constraint_force_before_ship_date_ignore_hinted(pre_solution)
        solver._add_constraint_sequence_subevents()
        # solver._add_constraint_machine_contiguous_block() # from what i can tell this isn't needed, i forget why but they already schedule contiguous

        solution = solver.solve()
        print(f"Solution status: {solution.status}, Objective value: {solution.objective_value}")



        if solution.status == "INFEASIBLE":
            print(f"Failed with a presolve range of {presolve_days} days. Trying again with larger presolve range.")
        else:
            print(f"Succeeded with a presolve range of {presolve_days} days.")
            if write_solution_to_excel:
                write_event_sequence_into_excel(solution, excel_path)
            break
    
    assert instance is not None, "Instance should have been created in the presolve loop."
    assert solution is not None, "Solver failed to find any solution across all presolve ranges."
    # save the final solution to csv for debugging organized by start date then machine
    # solution_df = pd.DataFrame(solution.schedule)
    # solution_df = solution_df.sort_values(by=["scheduledStartDate", "assignedMachineId"])
    # solution_df.to_csv("Inputs/final_solution.csv", index=False)

    workday_minutes = 8 * 60

    def _model_minutes_to_excel_date(model_minutes: int) -> date:
        clamped_minutes = max(0, int(model_minutes))
        day_offset = clamped_minutes // workday_minutes
        actual_date = date.today() + timedelta(days=day_offset)
        return actual_date

    # Graph view
    if solution.status in ["OPTIMAL", "FEASIBLE"]:
        schedules_to_plot = [solution.schedule, *solution.equally_optimal_schedules]
        total_solutions = len(schedules_to_plot)
        if total_solutions > 1:
            print(f"Rendering {total_solutions} equally optimal schedule graphs.")

        for solution_index, schedule_to_plot in enumerate(schedules_to_plot, start=1):
            title = f"Schedule by Machine (objective {solution.objective_value:g})"
            if total_solutions > 1:
                title = (
                    f"Schedule by Machine (solution {solution_index}/{total_solutions}, "
                    f"objective {solution.objective_value:g})"
                )

            interactive_output_path = None
            if total_solutions == 1:
                interactive_output_path = Path("Outputs") / "schedule_graph_interactive.html"

            _plot_schedule_graph(
                schedule_to_plot,
                instance,
                workday_minutes,
                title,
                interactive_output_path if save_graph else None,
            )

        if show_graph:
            plt.show()

if __name__ == "__main__":
    main(
        show_graph= input("Show graph? (y/n): ").strip().lower() == "y", 
        save_graph=input("Save graph as html? (y/n): ").strip().lower() == "y", 
        write_solution_to_excel=input("Write solution to Excel? (y/n): ").strip().lower() == "y"
    )